from functools import lru_cache
from pathlib import Path

import pandas as pd
import keras
import tensorflow as tf
from keras.utils import FeatureSpace
from openpyxl import load_workbook, Workbook


import os
os.environ["KERAS_BACKEND"] = "tensorflow"
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'


__all__ = (
    'load_info',
    'drop_other_diagnoses',
    'diagnoses_to_digits',
    'dataframe_to_dataset',
    'prepare_datasets',
    'build_feature_space',
    'build_model',
)


def build_feature_space():
    # TODO excluded columns
    data: pd.DataFrame = load_info()
    diagnoses_to_digits(data)
    ds_train, ds_validation = prepare_datasets(data, sample_fraction=0.1)
    features = FeatureSpace(
        features={
            'sex'   : FeatureSpace.integer_categorical(),
            'age'   : FeatureSpace.float_rescaled(),
            'IN T'  : FeatureSpace.float_normalized(),
            'IN P'  : FeatureSpace.float_normalized(),
            'IN P1' : FeatureSpace.float_normalized(),
            'IN P2' : FeatureSpace.float_normalized(),
            'IN P3' : FeatureSpace.float_normalized(),
            'OUT T' : FeatureSpace.float_normalized(),
            'OUT P' : FeatureSpace.float_normalized(),
            'OUT P1': FeatureSpace.float_normalized(),
            'OUT P2': FeatureSpace.float_normalized(),
            'OUT P3': FeatureSpace.float_normalized(),
        },
        output_mode='concat'
    )
    train_ds_with_no_labels = ds_train.map(lambda x, _: x)
    features.adapt(train_ds_with_no_labels)
    return features


def drop_other_diagnoses(
    data: pd.DataFrame,
    exclude_columns: list[str] = None
) -> pd.DataFrame:
    indexes = []
    for index, row in enumerate(data['diagnose']):
        if row not in {'норма', 'астма ремиссия'}:
            indexes.append(index)
    data.drop(indexes, inplace=True)

    if exclude_columns is not None:
        data.drop(columns=exclude_columns, inplace=True)

    feature_space = build_feature_space()
    features = set(feature_space.features)
    features.add('diagnose')
    features = set(data.columns).difference(features)
    if features:
        data.drop(columns=list(features), inplace=True)

    return data


def diagnoses_to_digits(df: pd.DataFrame) -> None:
    diagnoses = df['diagnose'].unique().tolist()
    assert 'норма' in diagnoses, f"There's no 'норма' in diagnoses ({diagnoses})"
    diagnoses.remove('норма')
    dictionary = {diagnose: index for index, diagnose in enumerate(diagnoses, start=1)}
    dictionary['норма'] = 0
    df.replace({'diagnose': dictionary}, inplace=True)


def dataframe_to_dataset(dataframe: pd.DataFrame) -> tf.data.Dataset:
    dataframe = dataframe.copy()
    labels = dataframe.pop("diagnose")
    ds = tf.data.Dataset.from_tensor_slices((dict(dataframe), labels))
    ds = ds.shuffle(buffer_size=len(dataframe))
    return ds


def prepare_datasets(
    data: pd.DataFrame,
    sample_fraction: float
) -> tuple[tf.data.Dataset, tf.data.Dataset]:
    """
    :param data: Pandas dataframe with data
    :param sample_fraction: Fraction of axis items to return. Cannot be used with n.
    :return: Train and validation dataset
    """
    df_validation = data.sample(frac=sample_fraction, random_state=1658179)
    df_train = data.drop(df_validation.index)
    ds_validation = dataframe_to_dataset(df_validation)
    ds_train = dataframe_to_dataset(df_train)
    ds_validation = ds_validation.batch(32)
    ds_train = ds_train.batch(32)
    return ds_train, ds_validation


def load_info() -> pd.DataFrame:
    wb: Workbook = load_workbook(filename='ПАТТЕРН ОБЕЗЛИЧ..xlsx')
    columns = [
        'sex', 'age',
        'IN T', 'IN P', 'IN P1/P', 'IN P2/P', 'IN P3/P', 'IN P1', 'IN P2', 'IN P3',
        'OUT T', 'OUT P', 'OUT P1/P', 'OUT P2/P', 'OUT P3/P', 'OUT P1', 'OUT P2', 'OUT P3',
        'SUM T', 'SUM P', 'SUM P1/P', 'SUM P2/P', 'SUM P3/P', 'SUM P1', 'SUM P2', 'SUM P3',
        'diagnose'
    ]
    iterator = wb.worksheets[0]
    iterator = (row for row in iterator.iter_rows(min_row=3, min_col=0, max_col=len(columns)))
    iterator = ([i.value for i in row] for row in iterator)
    df = pd.DataFrame(iterator, columns=columns)
    return df


def load_all(path: Path, name: str) -> tuple[keras.Model, keras.utils.FeatureSpace]:
    path_model: Path = path / f"{name}.keras"
    path_featurespace = path / f"{name}_featurespace.keras"

    assert path_model.exists(), f"Can't find model in {path_model}"
    assert path_featurespace.exists(), f"Can't find FeatureSpace in {path_featurespace}"

    model: keras.Model = keras.saving.load_model(path_model)
    feature_space: keras.utils.FeatureSpace = keras.saving.load_model(path_featurespace)
    return model, feature_space


def predict(
    model: keras.Model,
    featurespace: keras.utils.FeatureSpace,
    parameters: dict[str, tf.Tensor] | dict[str, float | int]
) -> float:
    assert isinstance(model, keras.Model)
    assert isinstance(featurespace, keras.utils.FeatureSpace)
    assert len(parameters) != 0, "Arguments amount can't be null"
    if not isinstance(parameters[str(next(iter(parameters.keys())))], tf.Tensor):
        parameters = {name: tf.convert_to_tensor([value]) for name, value in parameters.items()}
    return model.predict({'input_1': featurespace(parameters)}, verbose=0)[0][0]


def build_model(encoded_features) -> keras.Model:
    x = keras.layers.Dense(32, activation="relu")(encoded_features)
    x = keras.layers.Dropout(0.2)(x)
    x = keras.layers.Dense(16, activation='relu')(x)
    x = keras.layers.Dropout(0.1)(x)
    predictions = keras.layers.Dense(1, activation="sigmoid")(x)

    training_model = keras.Model(inputs=encoded_features, outputs=predictions)
    training_model.compile(
        optimizer="adam", loss="binary_crossentropy", metrics=["accuracy"]
    )

    return training_model


def main():
    model, feature_space = load_all(Path('models'), 'Test')
    values_ok = {
        'sex': 1,
        'age': 16,
        'IN T': 1.44,
        'IN P': 18.10,
        'IN P1': 10.1,
        'IN P2': 4.7,
        'IN P3': 3.1,
        'OUT T': 0.8,
        'OUT P': 32.4,
        'OUT P1': 23.8,
        'OUT P2': 6.1,
        'OUT P3': 2.4
    }
    values_remission = {
        'sex': 1,
        'age': 17,
        'IN T': 3.32,
        'IN P': 86.956,
        'IN P1': 61.432,
        'IN P2': 20.131,
        'IN P3': 5.393,
        'OUT T': 2.4,
        'OUT P': 37.61,
        'OUT P1': 24.445,
        'OUT P2': 9.94,
        'OUT P3': 3.224
    }
    print(predict(model, feature_space, **values_ok))
    print(predict(model, feature_space, **values_remission))
